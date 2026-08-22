import json
import threading
from pathlib import Path
from typing import Any, Dict, List, Optional
from pydantic import BaseModel, Field
import pymysql
import pymysql.cursors

ROOT = Path(__file__).resolve().parents[1]
STORE = ROOT / "data" / "datasources.json"


class DataSourceConfig(BaseModel):
    id: str
    name: str
    type: str  # doris, mysql, postgresql, api, excel
    host: str = ""
    port: int = 0
    user: str = ""
    password: str = ""
    database: Optional[str] = None
    enabled: bool = True
    extra: Dict[str, Any] = Field(default_factory=dict)
    # API-specific
    api_url: Optional[str] = None
    api_method: str = "GET"
    api_headers: Dict[str, str] = Field(default_factory=dict)
    # Excel/CSV-specific
    file_path: Optional[str] = None


class DataSourceStore:
    def __init__(self):
        self._lock = threading.Lock()
        STORE.parent.mkdir(parents=True, exist_ok=True)

    def _load(self) -> Dict[str, dict]:
        if not STORE.exists():
            return {}
        try:
            return json.loads(STORE.read_text(encoding="utf-8"))
        except Exception:
            return {}

    def list(self) -> List[dict]:
        return list(self._load().values())

    def get(self, ds_id: str) -> Optional[dict]:
        return self._load().get(ds_id)

    def save(self, cfg: DataSourceConfig) -> dict:
        with self._lock:
            data = self._load()
            data[cfg.id] = cfg.model_dump()
            STORE.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
            return data[cfg.id]

    def delete(self, ds_id: str):
        with self._lock:
            data = self._load()
            if ds_id not in data:
                raise KeyError(ds_id)
            del data[ds_id]
            STORE.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")

    def test_connection(self, cfg: DataSourceConfig) -> Dict[str, Any]:
        """Test connectivity and return basic info."""
        try:
            if cfg.type == "api":
                return self._test_api(cfg)
            elif cfg.type == "excel":
                return self._test_excel(cfg)
            conn = self._connect(cfg)
            conn.close()
            return {"success": True, "message": "连接成功"}
        except Exception as e:
            return {"success": False, "message": str(e)}

    def _test_api(self, cfg: DataSourceConfig) -> Dict[str, Any]:
        import urllib.request
        url = cfg.api_url
        if not url:
            return {"success": False, "message": "API URL 未配置"}
        req = urllib.request.Request(url, method="HEAD")
        for k, v in cfg.api_headers.items():
            req.add_header(k, v)
        try:
            with urllib.request.urlopen(req, timeout=10) as resp:
                return {"success": True, "message": f"HTTP {resp.status} - 连接成功"}
        except Exception as e:
            # HEAD might not be supported, try GET
            try:
                req2 = urllib.request.Request(url, method="GET")
                for k, v in cfg.api_headers.items():
                    req2.add_header(k, v)
                with urllib.request.urlopen(req2, timeout=10) as resp:
                    return {"success": True, "message": f"HTTP {resp.status} - 连接成功"}
            except Exception as e2:
                return {"success": False, "message": str(e2)}

    def _test_excel(self, cfg: DataSourceConfig) -> Dict[str, Any]:
        fp = cfg.file_path
        if not fp:
            return {"success": False, "message": "文件路径未配置"}
        p = Path(fp)
        if not p.exists():
            # Check in data/uploads
            p = ROOT / "data" / "uploads" / fp
        if not p.exists():
            return {"success": False, "message": f"文件不存在: {fp}"}
        return {"success": True, "message": f"文件存在: {p.name} ({p.stat().st_size} bytes)"}

    def _connect(self, cfg: DataSourceConfig):
        if cfg.type in ("doris", "mysql"):
            return pymysql.connect(
                host=cfg.host,
                port=cfg.port,
                user=cfg.user,
                password=cfg.password,
                database=cfg.database or None,
                cursorclass=pymysql.cursors.DictCursor,
                connect_timeout=5,
                read_timeout=15,
            )
        elif cfg.type == "postgresql":
            try:
                import psycopg2
                import psycopg2.extras
                return psycopg2.connect(
                    host=cfg.host,
                    port=cfg.port,
                    user=cfg.user,
                    password=cfg.password,
                    dbname=cfg.database or "postgres",
                    connect_timeout=5,
                )
            except ImportError:
                raise RuntimeError("psycopg2 未安装，请 pip install psycopg2-binary")
        else:
            raise ValueError(f"不支持的数据源类型: {cfg.type}")

    def scan_metadata(self, cfg: DataSourceConfig):
        """Scan metadata from a configured datasource."""
        from .models import MetadataCatalog, MetadataDatabase, MetadataTable, MetadataColumn, MetadataSnapshot

        if cfg.type == "doris":
            return self._scan_doris(cfg)
        elif cfg.type == "mysql":
            return self._scan_mysql(cfg)
        elif cfg.type == "postgresql":
            return self._scan_postgresql(cfg)
        elif cfg.type == "api":
            return self._scan_api(cfg)
        elif cfg.type == "excel":
            return self._scan_excel(cfg)
        else:
            raise ValueError(f"不支持的数据源类型: {cfg.type}")

    def _scan_doris(self, cfg: DataSourceConfig):
        from .models import MetadataCatalog, MetadataDatabase, MetadataTable, MetadataColumn, MetadataSnapshot

        conn = self._connect(cfg)
        catalogs = []
        try:
            with conn.cursor() as cur:
                cur.execute("SHOW CATALOGS")
                for crow in cur.fetchall():
                    name = str(crow.get("CatalogName") or crow.get("Catalog") or next(iter(crow.values())))
                    ctype = str(crow.get("Type") or crow.get("type") or "unknown")
                    databases = []
                    try:
                        cur.execute(f"SHOW DATABASES FROM `{name}`")
                        for drow in cur.fetchall():
                            db = str(next(iter(drow.values())))
                            if db.lower() in {"information_schema", "mysql"}:
                                continue
                            tables = self._scan_doris_tables(cur, name, db)
                            databases.append(MetadataDatabase(name=db, tables=tables))
                    except Exception:
                        databases = []
                    catalogs.append(MetadataCatalog(name=name, type=ctype, databases=databases))
        finally:
            conn.close()
        return MetadataSnapshot(source=f"doris:{cfg.name}", catalogs=catalogs)

    def _scan_doris_tables(self, cur, catalog, db):
        from .models import MetadataTable, MetadataColumn
        tables = []
        try:
            cur.execute(f"SHOW TABLES FROM `{catalog}`.`{db}`")
            for trow in cur.fetchall():
                tname = str(next(iter(trow.values())))
                columns = []
                try:
                    cur.execute(f"DESC `{catalog}`.`{db}`.`{tname}`")
                    for col in cur.fetchall():
                        columns.append(MetadataColumn(
                            name=str(col.get("Field") or next(iter(col.values()))),
                            data_type=str(col.get("Type") or "UNKNOWN"),
                            nullable=str(col.get("Null") or "YES").upper() != "NO",
                        ))
                except Exception:
                    pass
                tables.append(MetadataTable(name=tname, columns=columns))
        except Exception:
            pass
        return tables

    def _scan_mysql(self, cfg: DataSourceConfig):
        from .models import MetadataCatalog, MetadataDatabase, MetadataTable, MetadataColumn, MetadataSnapshot

        conn = self._connect(cfg)
        databases = []
        try:
            with conn.cursor() as cur:
                if cfg.database:
                    db_names = [cfg.database]
                else:
                    cur.execute("SHOW DATABASES")
                    db_names = [str(next(iter(r.values()))) for r in cur.fetchall()
                                if str(next(iter(r.values()))).lower() not in {"information_schema", "mysql", "performance_schema", "sys"}]
                for db in db_names:
                    cur.execute(f"USE `{db}`")
                    cur.execute("SHOW TABLES")
                    tables = []
                    for trow in cur.fetchall():
                        tname = str(next(iter(trow.values())))
                        cur.execute(f"DESC `{tname}`")
                        columns = []
                        for col in cur.fetchall():
                            columns.append(MetadataColumn(
                                name=str(col.get("Field") or next(iter(col.values()))),
                                data_type=str(col.get("Type") or "UNKNOWN"),
                                nullable=str(col.get("Null") or "YES").upper() != "NO",
                            ))
                        tables.append(MetadataTable(name=tname, columns=columns))
                    databases.append(MetadataDatabase(name=db, tables=tables))
        finally:
            conn.close()
        catalog = MetadataCatalog(name=cfg.name, type="mysql", databases=databases)
        return MetadataSnapshot(source=f"mysql:{cfg.name}", catalogs=[catalog])

    def _scan_postgresql(self, cfg: DataSourceConfig):
        from .models import MetadataCatalog, MetadataDatabase, MetadataTable, MetadataColumn, MetadataSnapshot

        conn = self._connect(cfg)
        tables = []
        try:
            cur = conn.cursor()
            cur.execute("""
                SELECT table_schema, table_name
                FROM information_schema.tables
                WHERE table_schema NOT IN ('pg_catalog', 'information_schema')
                  AND table_type = 'BASE TABLE'
                ORDER BY table_schema, table_name
            """)
            for schema_name, table_name in cur.fetchall():
                cur.execute("""
                    SELECT column_name, data_type, is_nullable
                    FROM information_schema.columns
                    WHERE table_schema = %s AND table_name = %s
                    ORDER BY ordinal_position
                """, (schema_name, table_name))
                columns = [
                    MetadataColumn(name=row[0], data_type=row[1], nullable=(row[2] == "YES"))
                    for row in cur.fetchall()
                ]
                tables.append(MetadataTable(name=f"{schema_name}.{table_name}", columns=columns))
            cur.close()
        finally:
            conn.close()
        db = MetadataDatabase(name=cfg.database or "postgres", tables=tables)
        catalog = MetadataCatalog(name=cfg.name, type="postgresql", databases=[db])
        return MetadataSnapshot(source=f"postgresql:{cfg.name}", catalogs=[catalog])

    def _scan_api(self, cfg: DataSourceConfig):
        """Scan an API endpoint, infer schema from JSON response."""
        import urllib.request
        from .models import MetadataCatalog, MetadataDatabase, MetadataTable, MetadataColumn, MetadataSnapshot

        url = cfg.api_url
        if not url:
            raise ValueError("API URL 未配置")
        req = urllib.request.Request(url, method=cfg.api_method)
        for k, v in cfg.api_headers.items():
            req.add_header(k, v)
        with urllib.request.urlopen(req, timeout=15) as resp:
            body = json.loads(resp.read().decode("utf-8"))

        # Infer columns from JSON response
        tables = []
        records = body if isinstance(body, list) else body.get("data", body.get("results", body.get("items", [body])))
        if not isinstance(records, list):
            records = [records]
        if records:
            sample = records[0] if isinstance(records[0], dict) else {}
            columns = []
            for key, val in sample.items():
                dtype = "VARCHAR"
                if isinstance(val, (int, float)):
                    dtype = "DOUBLE" if isinstance(val, float) else "BIGINT"
                elif isinstance(val, bool):
                    dtype = "BOOLEAN"
                columns.append(MetadataColumn(name=key, data_type=dtype))
            endpoint_name = url.rstrip("/").split("/")[-1] or "api_data"
            tables.append(MetadataTable(name=endpoint_name, columns=columns))

        db = MetadataDatabase(name="api", tables=tables)
        catalog = MetadataCatalog(name=cfg.name, type="api", databases=[db])
        return MetadataSnapshot(source=f"api:{cfg.name}", catalogs=[catalog])

    def _scan_excel(self, cfg: DataSourceConfig):
        """Scan an Excel/CSV file, infer schema from headers and data types."""
        import csv
        from .models import MetadataCatalog, MetadataDatabase, MetadataTable, MetadataColumn, MetadataSnapshot

        fp = cfg.file_path
        if not fp:
            raise ValueError("文件路径未配置")
        p = Path(fp)
        if not p.exists():
            p = ROOT / "data" / "uploads" / fp
        if not p.exists():
            raise ValueError(f"文件不存在: {fp}")

        tables = []
        suffix = p.suffix.lower()

        if suffix == ".csv":
            tables.append(self._scan_csv_file(p))
        elif suffix in (".xlsx", ".xls"):
            tables.extend(self._scan_excel_file(p))
        else:
            raise ValueError(f"不支持的文件格式: {suffix}，仅支持 .csv / .xlsx / .xls")

        db = MetadataDatabase(name="file", tables=tables)
        catalog = MetadataCatalog(name=cfg.name, type="excel", databases=[db])
        return MetadataSnapshot(source=f"excel:{cfg.name}", catalogs=[catalog])

    def _scan_csv_file(self, path: Path):
        import csv
        from .models import MetadataTable, MetadataColumn

        with open(path, "r", encoding="utf-8-sig") as f:
            reader = csv.DictReader(f)
            headers = reader.fieldnames or []
            # Read first few rows to infer types
            rows = []
            for i, row in enumerate(reader):
                if i >= 20:
                    break
                rows.append(row)

        columns = []
        for h in headers:
            dtype = self._infer_csv_type([r.get(h, "") for r in rows])
            columns.append(MetadataColumn(name=h, data_type=dtype))
        return MetadataTable(name=path.stem, columns=columns)

    def _scan_excel_file(self, path: Path):
        from .models import MetadataTable, MetadataColumn
        try:
            import openpyxl
        except ImportError:
            raise RuntimeError("openpyxl 未安装，请 pip install openpyxl")

        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        tables = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows_iter = ws.iter_rows(values_only=True)
            headers = next(rows_iter, None)
            if not headers:
                continue
            headers = [str(h or f"col_{i}") for i, h in enumerate(headers)]
            sample_rows = []
            for i, row in enumerate(rows_iter):
                if i >= 20:
                    break
                sample_rows.append(row)

            columns = []
            for idx, h in enumerate(headers):
                values = [r[idx] if idx < len(r) else None for r in sample_rows]
                dtype = self._infer_excel_type(values)
                columns.append(MetadataColumn(name=h, data_type=dtype))
            tables.append(MetadataTable(name=sheet_name, columns=columns))
        wb.close()
        return tables

    @staticmethod
    def _infer_csv_type(values: List[str]) -> str:
        """Infer column type from string values."""
        non_empty = [v for v in values if v.strip()]
        if not non_empty:
            return "VARCHAR"
        numeric = 0
        for v in non_empty:
            try:
                float(v.replace(",", ""))
                numeric += 1
            except ValueError:
                pass
        if numeric > len(non_empty) * 0.8:
            if all("." not in v and "e" not in v.lower() for v in non_empty if v.strip()):
                return "BIGINT"
            return "DOUBLE"
        # Check datetime-like
        import re
        date_pattern = re.compile(r"\d{4}[-/]\d{1,2}[-/]\d{1,2}")
        if sum(1 for v in non_empty if date_pattern.search(v)) > len(non_empty) * 0.5:
            return "DATETIME"
        return "VARCHAR"

    @staticmethod
    def _infer_excel_type(values: list) -> str:
        """Infer column type from Python values."""
        from datetime import datetime
        non_none = [v for v in values if v is not None]
        if not non_none:
            return "VARCHAR"
        types = set(type(v).__name__ for v in non_none)
        if types <= {"int"}:
            return "BIGINT"
        if types <= {"int", "float"}:
            return "DOUBLE"
        if types <= {"datetime"}:
            return "DATETIME"
        if any(isinstance(v, datetime) for v in non_none):
            return "DATETIME"
        return "VARCHAR"
